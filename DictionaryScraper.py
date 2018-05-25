'''
By Austin Dorsey
Started: 4/9/18
Last modified 4/15/18
To soothe my curiousity of what is the most used word used to descride another 
word in the dictionary. Dictionary used is dictionary.com
'''

from queue import Queue
from threading import Thread
import time
from openpyxl import Workbook
from lxml import html
import requests

def read_dictionary(dictionary, max_words = 0, start_letter = 'a', end_letter = 'z'):
    '''Gets all the listed words in the letter range upto the number of max_words.'''
    num_words = 0
    for l in range(ord(start_letter), ord(end_letter) + 1):
        first_word = ""
        x = 0
        while True:
            x += 1 
            print(chr(l) + ' - ' + str(x))
            url = 'http://www.dictionary.com/list/' + chr(l) + '/' + str(x)
            try:
                page = requests.get(url, timeout=30)
            except:
                print('\tURL request failed, retrying.')
                x -= 1
                continue
            tree = html.document_fromstring(page.content)
            #Path for all the words on that page.
            temp_words = tree.xpath('/html/body/div[2]/div[3]/ul/li/span/a/text()')
            if x != 1 and temp_words[0] == first_word:
                break
            else:
                if x == 1:
                    first_word = temp_words[0]
                for word in temp_words:
                    if num_words >= max_words and max_words != 0:
                        return
                    if word.find(' ') != -1 or word.find('.') != -1:
                        continue
                    num_words += 1
                    dictionary[word.lower()] = 0



def read_word(word, invalad_URL):
    '''Returns the description of the geven word.'''
    try:
        print(word)
        url = 'http://www.dictionary.com/browse/' + word
        page = requests.get(url, timeout=30)
        tree = html.document_fromstring(page.content)
        #Path for the deffinitions of the word.
        raw_discript = tree.xpath('/html/body/div[4]/div[1]/div[2]/section[1]/div[1]/section/div/div[1]/section/div/div//text()')
        return raw_discript
    except:
        print('\tinvaled url or timeout:', url)
        invalad_URL.append(word)
        return None


def raw_process(raw, dictionary, words_not_in_dictionary):
    '''Takes the raw list of the discription, and prosesses it down to the word and counts it.'''
    for x in raw:
        discript = x.split(" ")
        for y in discript:
            if not y.isalpha():
                y = y.replace('.', '')
                y = y.replace(',', '')
                y = y.replace(':', '')
                y = y.replace(';', '')
                y = y.replace('(', '')
                y = y.replace(')', '')
                y = y.replace('\n', '')
                y = y.replace('\t', '')
                y = y.lstrip()
            if y == "" or y.isnumeric():
                continue
            y = y.lower()

            try:
                dictionary[y] += 1
            except KeyError:
                words_not_in_dictionary.add(y)


class ReadWorker(Thread):
    '''Takes words in the dictionary and requests their dicriptions.
    Dicriptions are then put into raw_queue.
    '''
    def __init__(self, queue, raw_queue, invalad_URL):
        Thread.__init__(self)
        self.queue = queue
        self.raw_queue = raw_queue
        self.invalad_URL = invalad_URL

    def run(self):
        while True:
            try:
                word = self.queue.get_nowait()
                self.raw_queue.put(read_word(word, self.invalad_URL))
                self.queue.task_done()
            except:
                break


class ProcessWorker(Thread):
    '''Takes raw discriptions and calls raw_process which counds the number
    of times that each word is used in the discriptions.
    '''
    def __init__(self, queue, dictionary, words_not_in_dictionary):
        Thread.__init__(self)
        self.queue = queue
        self.dictionary = dictionary
        self.words_not_in_dictionary = words_not_in_dictionary

    def run(self):
        while True:
            raw = self.queue.get()
            if raw is None:
                self.queue.task_done()
                continue
            raw_process(raw, self.dictionary, self.words_not_in_dictionary)
            self.queue.task_done()


def main():
    '''Gets the list of words on dictionary.com, creates threads to get the
    discription of those words, counts the number of times a word is used
    in the discriptions, and saves the resaults to an Excel spread sheet.
    '''
    dictionary = {}
    invalad_URL = []
    words_not_in_dictionary = set()

    ts = time.clock()
    read_dictionary(dictionary, 100, 'a', 'z')
    print('Ending reading dictionary.com', len(dictionary), 'words.')
    print('Time taken', time.clock() - ts)

    ts = time.clock()

    #Queue
    queue = Queue()
    raw_queue = Queue()
    print('Filling queue.')
    for word in dictionary.keys():
        if word.find(' ') != -1 or word.find('.') != -1:
            continue
        queue.put(word)

    #Threads
    workers = []
    print('Creating workers.')
    for x in range(36):
        worker = ReadWorker(queue, raw_queue, invalad_URL)
        worker.daemon = True
        worker.start()
        workers.append(worker)

    for x in range(1):
        process_worker = ProcessWorker(raw_queue, dictionary, words_not_in_dictionary)
        process_worker.daemon = True
        process_worker.start()

    #Checking if done.
    try:
        num_stopped = 0
        while num_stopped < 36:
            for i, worker in enumerate(workers):
                worker.join(1)
                if not worker.is_alive():
                    num_stopped += 1
                    del workers[i]
                    print(num_stopped)
                    break
        print('Finished reading words', time.clock() - ts)
        try:
            while True:
                raw_process(raw_queue.get_nowait(), dictionary, words_not_in_dictionary)
                raw_queue.task_done()
        except:
            print('Finished processing')
    except KeyboardInterrupt:
        with queue.mutex:
            queue.queue.clear()
        print('Interupted word reading.')

    #Setting up workbook.
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Resaults"
    ws2 = wb.create_sheet(title="!inDictionary")
    ws3 = wb.create_sheet(title="Invalad URL")

    #Adding to excel and printing dictionary.
    for i, n in enumerate(dictionary):
        ws1['A' + str(i+1)] = n
        ws1['B' + str(i+1)] = dictionary[n]
    wb.save(filename='Resaults.xlsx')

    #Sorts the data by value. Highest value gets printed last.
    sorted_key = sorted(dictionary, key=dictionary.get, reverse=False)
    for n in sorted_key:
        if dictionary[n] != 0:
            print(n, dictionary[n])

    #Words found in the descriptions that were not found in the dictionary.
    print(words_not_in_dictionary)
    for i, word in enumerate(words_not_in_dictionary):
        ws2['A' + str(i+1)] = word
    
    #Showing invalad URLs
    for i, url in enumerate(invalad_URL):
        ws3['A' + str(i+1)] = url
        print(url)

    #Saving excel sheet.
    wb.save(filename='Resaults.xlsx')


if __name__ == '__main__':
    main()
